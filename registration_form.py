from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter.simpledialog import askstring
from tabulate import tabulate

file = load_workbook("C:\\Users\\Admin\\Desktop\\Proiect PYTHON facultate\\data.xlsx")

sheet = file.active

def excel():
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    
    sheet.cell(row=1, column=1).value = "Nume"
    sheet.cell(row=1, column=2).value = "Facultate"
    sheet.cell(row=1, column=3).value = "Serie"
    sheet.cell(row=1, column=4).value = "Grupa"
    sheet.cell(row=1, column=5).value = "Email"
    sheet.cell(row=1, column=6).value = "Adresa domiciliu"
    sheet.cell(row=1, column=7).value = "Nr. de telefon"
    sheet.cell(row=1, column=8).value = "Data de nastere"
    
def focus1(event):
    facultate_spatiu.focus_set()
    
def focus2(event):
    serie_spatiu.focus_set()
        
def focus3(event):
    grupa_spatiu.focus_set()
        
def focus4(event):
    email_spatiu.focus_set()
        
def focus5(event):
    adresa_spatiu.focus_set()
        
def focus6(event):
    telefon_spatiu.focus_set()
    
def focus7(event):
    nastere_spatiu.focus_set()
        
def clear():
    nume_spatiu.delete(0, END)
    facultate_spatiu.delete(0, END)
    serie_spatiu.delete(0, END)
    grupa_spatiu.delete(0, END)
    email_spatiu.delete(0, END)
    adresa_spatiu.delete(0, END)
    telefon_spatiu.delete(0, END)
    nastere_spatiu.delete(0, END)
        
def insert():
    
    if(nume_spatiu.get() == "" and facultate_spatiu.get() == "" and 
       serie_spatiu.get() == "" and grupa_spatiu.get() == "" and 
       email_spatiu.get() == "" and adresa_spatiu.get() == "" and
       telefon_spatiu.get() == "" and nastere_spatiu.get() == ""):
        messagebox.showerror("Eroare", "Completati toate campurile!")
    
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        
        sheet.cell(row = current_row + 1, column= 1).value = nume_spatiu.get()
        sheet.cell(row = current_row + 1, column= 2).value = facultate_spatiu.get()
        sheet.cell(row = current_row + 1, column= 3).value = serie_spatiu.get()
        sheet.cell(row = current_row + 1, column= 4).value = grupa_spatiu.get()
        sheet.cell(row = current_row + 1, column= 5).value = email_spatiu.get()
        sheet.cell(row = current_row + 1, column= 6).value = adresa_spatiu.get()
        sheet.cell(row = current_row + 1, column= 7).value = telefon_spatiu.get()
        sheet.cell(row = current_row + 1, column= 8).value = nastere_spatiu.get()
        
        file.save("C:\\Users\\Admin\\Desktop\\Proiect PYTHON facultate\\data.xlsx")

        nume_spatiu.focus_set()
        
        clear()

def view_people():
    view_people = Toplevel(window)
    view_people.title("Inregistrari")
    view_people.geometry("800x400")
    
    text_area = Text(view_people, wrap=NONE)
    text_area.pack(expand=1, fill=BOTH)
    
    persoane = list(sheet.iter_rows(values_only=True))
    if not persoane:
        text_area.insert(1.0, "Nimic aici")
        return
    
    tabel = tabulate(persoane, tablefmt="grid")
    
    text_area.insert(1.0, tabel)
    text_area.config(state=DISABLED)

def reset():
    password = askstring("Aceasta optiune este rezervata adminului", "Introduceti parola")
    if password != "Sendvis_Popeyes":
        messagebox.showerror("Eroare", "Parola incorecta!")
        return
    
    confirmare = messagebox.askyesno("Confirmare?", "Actiunea aceasta va sterge toate datele din baza")
    if not confirmare:
        return
    
    categorii = [cell.value for cell in sheet[1]]
    
    sheet.delete_rows(2, sheet.max_row)
    
    for col_num, value in enumerate(categorii, start=1):
        sheet.cell(row=1, column=col_num).value = value
    
    file.save("data.xlsx")
    messagebox.showinfo("Succes", "Toate datele au fost sterse!")
    
if __name__ == "__main__":
    
    window = Tk()
    
    window.configure(background="yellow")
    
    window.title("Inregistrare elev la UPB")
    
    window.geometry("800x400")
    
    excel()
    
    window.columnconfigure(1, weight=1)
    
    heading = Label(window, text="Student UPB", bg="yellow", font=("Comic Sans", 18))
    nume = Label(window, text="Nume", bg="yellow")
    facultate = Label(window, text="Facultate", bg="yellow")
    serie = Label(window, text="Serie", bg="yellow")
    grupa = Label(window, text="Grupa", bg="yellow")
    email = Label(window, text="Email", bg="yellow")
    adresa = Label(window, text="Adresa domiciliu", bg="yellow")
    telefon = Label(window, text="Nr. de telefon", bg="yellow")
    nastere = Label(window, text="Data de nastere", bg="yellow")
    
    heading.grid(row=0, column=1, sticky = "ew")
    nume.grid(row=1, column=0, sticky = "w")
    facultate.grid(row=2, column=0, sticky = "w")
    serie.grid(row=3, column=0, sticky = "w")
    grupa.grid(row=4, column=0, sticky = "w")
    email.grid(row=5, column=0, sticky = "w")
    adresa.grid(row=6, column=0, sticky = "w")
    telefon.grid(row=7, column=0, sticky = "w")
    nastere.grid(row=8, column=0, sticky = "w")
    
    nume_spatiu = Entry(window)
    facultate_spatiu = Entry(window)
    serie_spatiu = Entry(window)
    grupa_spatiu = Entry(window)
    email_spatiu = Entry(window)
    adresa_spatiu = Entry(window)
    telefon_spatiu = Entry(window)
    nastere_spatiu = Entry(window)
    
    nume_spatiu.bind("<Return>", focus1)
    facultate_spatiu.bind("<Return>", focus2)
    serie_spatiu.bind("<Return>", focus3)
    grupa_spatiu.bind("<Return>", focus4)
    email_spatiu.bind("<Return>", focus5)
    adresa_spatiu.bind("<Return>", focus6)
    telefon_spatiu.bind("<Return>", focus7)
    
    nume_spatiu.grid(row=1, column=1, ipadx="250")
    facultate_spatiu.grid(row=2, column=1, ipadx="250")
    serie_spatiu.grid(row=3, column=1, ipadx="250")
    grupa_spatiu.grid(row=4, column=1, ipadx="250")
    email_spatiu.grid(row=5, column=1, ipadx="250")
    adresa_spatiu.grid(row=6, column=1, ipadx="250")
    telefon_spatiu.grid(row=7, column=1, ipadx="250")
    nastere_spatiu.grid(row=8, column=1, ipadx="250")
    
    excel()
    
    submit = Button(window, text= "Submit", fg="black", bg="green", command=insert)
    submit.grid(row = 9, column=1, pady= 10)
    
    view = Button(window, text="Inregistrari curente", fg="black", bg ="blue", command=view_people)
    view.grid(row=10, column=1, pady=10)
    
    reset_button = Button(window, text="Reset", fg = "black", bg= "red", command=reset)
    reset_button.grid(row = 11, column=1, pady=10)
    window.mainloop()